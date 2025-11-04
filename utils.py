"""
Утилиты для обработки данных Scopus
"""
import pandas as pd
import re
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime


def normalize_title(title):
    """
    Нормализует название статьи для сравнения.
    Убирает лишние пробелы, приводит к нижнему регистру.
    """
    if pd.isna(title):
        return ""
    return re.sub(r'\s+', ' ', str(title).lower().strip())


def find_new_articles(df_source, df_existing, threshold=95):
    """
    Находит статьи из df_source, которых нет в df_existing.

    Параметры:
    - df_source: DataFrame с исходными статьями (Scopus)
    - df_existing: DataFrame с существующими статьями (United)
    - threshold: порог fuzzy matching (0-100)

    Возвращает:
    - DataFrame с новыми статьями
    - dict с информацией о найденных дубликатах
    """
    # Нормализуем названия в существующих статьях
    existing_titles = df_existing['Title'].apply(normalize_title).tolist()

    new_articles = []
    duplicates_info = []

    for idx, row in df_source.iterrows():
        source_title = normalize_title(row['Title'])

        # Проверяем на дубликаты с fuzzy matching
        is_duplicate = False
        best_match_score = 0
        best_match_title = ""

        for existing_title in existing_titles:
            similarity = fuzz.ratio(source_title, existing_title)

            if similarity > best_match_score:
                best_match_score = similarity
                best_match_title = existing_title

            if similarity >= threshold:
                is_duplicate = True
                duplicates_info.append({
                    'source_title': row['Title'],
                    'matched_title': existing_title,
                    'similarity': similarity
                })
                break

        # Если не дубликат, добавляем в список новых
        if not is_duplicate:
            new_articles.append(idx)

    df_new = df_source.loc[new_articles].copy()

    return df_new, duplicates_info


def extract_affiliation_authors(authors_with_affiliations, author_full_names, keywords):
    """
    Извлекает авторов из строки с аффилиациями, в именах которых есть слова из `keywords`.

    Параметры:
    - authors_with_affiliations: строка "LastName, FirstName, affiliation; LastName, FirstName, affiliation; ..."
    - author_full_names: строка с полными именами и ID "Name (ID); Name (ID); ..."
    - keywords: список ключевых слов для поиска университета

    Возвращает:
    - dict с информацией об авторах искомых аффиляций:
        {
            'authors_short': 'LastName, F.; LastName2, F.',  # Краткие имена
            'authors_with_ids': 'Full Name (ID); Full Name2 (ID)',  # С ID
            'authors_full': 'Full Name; Full Name2',  # Полные имена
            'count': 2  # Количество авторов
        }
    """
    if pd.isna(authors_with_affiliations):
        return {'authors_short': '', 'authors_with_ids': '', 'authors_full': '', 'count': 0}

    # Разбиваем на отдельных авторов (разделитель - точка с запятой)
    author_blocks = str(authors_with_affiliations).split(';')

    affiliate_authors_short = []
    affiliate_authors_with_ids = []
    affiliate_authors_full = []

    # Парсим author_full_names для получения ID
    full_names_dict = {}
    if not pd.isna(author_full_names):
        # Формат: "LastName, FirstName (ID); LastName, FirstName (ID); ..."
        full_name_parts = str(author_full_names).split(';')
        for part in full_name_parts:
            part = part.strip()
            # Извлекаем имя и ID
            match = re.match(r'(.+?)\s*\((\d+)\)', part)
            if match:
                name = match.group(1).strip()
                author_id = match.group(2).strip()
                # Создаем ключ по фамилии для поиска
                name_parts = name.split(',')
                if len(name_parts) >= 1:
                    last_name = name_parts[0].strip()
                    full_names_dict[last_name] = {'full': name, 'id': author_id}

    # Проверяем каждого автора
    for block in author_blocks:
        block = block.strip()
        if not block:
            continue

        # Проверяем, содержит ли блок ключевые слова нужной аффиляции
        is_affiliated = any(keyword.lower() in block.lower() for keyword in keywords)

        if is_affiliated:
            # Извлекаем имя автора (формат: "LastName, FirstName, affiliation, affiliation, ...")
            # Имя автора - это первые две части до запятых
            parts = block.split(',')
            if len(parts) >= 2:
                last_name = parts[0].strip()
                first_name = parts[1].strip()

                # Краткое имя: "LastName, F."
                first_initial = first_name[0] + '.' if first_name else ''
                short_name = f"{last_name}, {first_initial}"
                affiliate_authors_short.append(short_name)

                # Полное имя с ID из author_full_names
                if last_name in full_names_dict:
                    full_info = full_names_dict[last_name]
                    affiliate_authors_with_ids.append(f"{full_info['full']} ({full_info['id']})")
                    affiliate_authors_full.append(full_info['full'])
                else:
                    # Если ID не найден, используем имя без ID
                    full_name = f"{last_name}, {first_name}"
                    affiliate_authors_with_ids.append(full_name)
                    affiliate_authors_full.append(full_name)

    return {
        'authors_short': '; '.join(affiliate_authors_short),
        'authors_with_ids': '; '.join(affiliate_authors_with_ids),
        'authors_full': '; '.join(affiliate_authors_full),
        'count': len(affiliate_authors_short)
    }


def map_departments(authors_string, df_dept_mapping):
    """
    Сопоставляет авторов с департаментами из справочника.

    Параметры:
    - authors_string: строка с авторами, разделенными "; " (например, "LastName, F.; LastName2, F.")
    - df_dept_mapping: DataFrame со справочником (колонки: 'Author Name', 'Departament')

    Возвращает:
    - dict с информацией:
        {
            'department': строка с департаментом(ами) через "; "
            'needs_highlight': bool - нужна ли желтая подсветка
            'reason': причина подсветки ('not_found', 'multiple', None)
        }
    """
    if pd.isna(authors_string) or not authors_string.strip():
        return {'department': '', 'needs_highlight': False, 'reason': None}

    # Разбиваем на отдельных авторов
    authors = [a.strip() for a in str(authors_string).split(';') if a.strip()]

    departments = []
    not_found_authors = []

    for author in authors:
        # Ищем автора в справочнике (нечувствительно к регистру)
        matches = df_dept_mapping[
            df_dept_mapping['Author Name'].str.lower() == author.lower()
        ]

        if len(matches) == 0:
            # Автор не найден
            not_found_authors.append(author)
        else:
            # Автор найден, извлекаем департамент(ы)
            for _, row in matches.iterrows():
                dept = row['Departament']
                if pd.notna(dept) and dept.strip():
                    departments.append(dept.strip())

    # Убираем дубликаты департаментов
    departments = list(dict.fromkeys(departments))  # Сохраняет порядок

    # Определяем нужна ли подсветка
    needs_highlight = False
    reason = None

    if not_found_authors:
        # Есть авторы без департаментов
        needs_highlight = True
        reason = 'not_found'
    elif len(departments) > 1:
        # Множественные департаменты
        needs_highlight = True
        reason = 'multiple'

    department_str = '; '.join(departments) if departments else ''

    return {
        'department': department_str,
        'needs_highlight': needs_highlight,
        'reason': reason,
        'not_found_authors': not_found_authors
    }


def process_scopus_data(df_source, df_existing, df_dept_mapping,
                       threshold=95, year=None, title_exclude_keywords=None,
                       affiliation_keywords=None):
    """
    Основной pipeline обработки данных Scopus.

    Параметры:
    - df_source: DataFrame с исходными данными Scopus
    - df_existing: DataFrame с существующими данными United
    - df_dept_mapping: DataFrame со справочником департаментов
    - threshold: порог fuzzy matching для поиска дубликатов
    - year: год(ы) для фильтрации статей
        None - без фильтрации (все годы)
        int - один год (например, 2025)
        list - несколько лет (например, [2024, 2025])
    - title_exclude_keywords: список подстрок для исключения статей по Title (case-insensitive)
    - affiliation_keywords: список ключевых слов для поиска авторов из аффилированных учреждений

    Возвращает:
    - df_result: DataFrame с новыми статьями в формате United
    - stats: словарь со статистикой обработки
    """

    stats = {
        'original_scopus_count': len(df_source),
        'original_united_count': len(df_existing),
        'after_year_filter_scopus': 0,
        'after_year_filter_united': 0,
        'after_title_filter': 0,
        'excluded_by_title': 0,
        'new_articles': 0,
        'duplicates_found': 0,
        'affiliated_articles': 0,
        'no_affiliated_authors': 0,
        'highlighted_depts': 0
    }

    # Фильтрация по году
    if year is not None:
        # Приводим год к списку для единообразной обработки
        years_list = [year] if isinstance(year, int) else year

        # Фильтруем Scopus
        df_source = df_source[df_source['Year'].isin(years_list)].copy()
        stats['after_year_filter_scopus'] = len(df_source)

        # Фильтруем United
        df_existing = df_existing[df_existing['Year'].isin(years_list)].copy()
        stats['after_year_filter_united'] = len(df_existing)
    else:
        stats['after_year_filter_scopus'] = len(df_source)
        stats['after_year_filter_united'] = len(df_existing)

    # Фильтрация по Title - исключение статей с определенными подстроками
    if title_exclude_keywords and len(title_exclude_keywords) > 0:
        before_filter_count = len(df_source)

        # Создаем маску: True если Title НЕ содержит ни одной из подстрок
        def should_keep_article(title):
            if pd.isna(title):
                return True
            title_lower = str(title).lower()
            # Если хотя бы одна подстрока найдена - исключаем (return False)
            for keyword in title_exclude_keywords:
                if keyword.lower() in title_lower:
                    return False
            return True

        df_source = df_source[df_source['Title'].apply(should_keep_article)].copy()

        stats['excluded_by_title'] = before_filter_count - len(df_source)
        stats['after_title_filter'] = len(df_source)
    else:
        stats['after_title_filter'] = len(df_source)

    # Шаг 1: Найти новые статьи
    df_new, duplicates = find_new_articles(df_source, df_existing, threshold)

    stats['new_articles'] = len(df_new)
    stats['duplicates_found'] = len(duplicates)

    if len(df_new) == 0:
        return pd.DataFrame(), stats

    # Шаг 2: Обработать каждую новую статью
    result_data = []

    for idx, (_, row) in enumerate(df_new.iterrows(), 1):
        # Извлечь авторов с нужной аффиляцией
        authors_info = extract_affiliation_authors(
            row['Authors with affiliations'],
            row['Author full names'],
            affiliation_keywords
        )

        # Пропустить статьи без авторов аффиляцией
        if authors_info['count'] == 0:
            stats['no_affiliated_authors'] += 1
            continue

        stats['affiliated_articles'] += 1

        # Сопоставить департаменты
        dept_info = map_departments(authors_info['authors_short'], df_dept_mapping)

        if dept_info['needs_highlight']:
            stats['highlighted_depts'] += 1

        # Сформировать запись
        result_row = {
            'Departament': dept_info['department'],
            'Authors': authors_info['authors_short'],  # Только авторы из нужных аффилиаций
            'Authors.1': row['Authors'] if 'Authors' in row else '',  # ВСЕ авторы
            'Author full names': row['Author full names'] if 'Author full names' in row else '',  # ВСЕ полные имена
            'Title': row['Title'],
            'Year': row['Year'] if 'Year' in row and pd.notna(row['Year']) else '',
            'Source title': row['Source title'] if 'Source title' in row else '',
            'Volume': row['Volume'] if 'Volume' in row else '',
            'Issue': row['Issue'] if 'Issue' in row else '',
            'Art. No.': row['Art. No.'] if 'Art. No.' in row else '',
            'Page start': row['Page start'] if 'Page start' in row else '',
            'Page end': row['Page end'] if 'Page end' in row else '',
            'Page count': row['Page count'] if 'Page count' in row else '',
            'Source': 'Scopus',
            'Təqdimat': '',
            'Data': '',
            'Amount': '',
            'Quartil': '',
            '_highlight': dept_info['needs_highlight'],
            '_highlight_reason': dept_info['reason']
        }

        result_data.append(result_row)

    # Создать итоговый DataFrame
    df_result = pd.DataFrame(result_data)

    return df_result, stats


def export_to_excel_with_highlighting(df, output_path, highlight_color='FFFF00'):
    """
    Экспортирует DataFrame в Excel с желтой подсветкой проблемных ячеек.

    Параметры:
    - df: DataFrame с данными (должен содержать колонки _highlight и _highlight_reason)
    - output_path: путь к выходному файлу
    - highlight_color: цвет подсветки в формате RGB hex (по умолчанию желтый)

    Возвращает:
    - True если успешно, False если ошибка
    """

    if len(df) == 0:
        return False

    try:
        # Создаем копию без служебных колонок
        df_export = df.drop(columns=['_highlight', '_highlight_reason'], errors='ignore').copy()

        # Экспортируем в Excel
        df_export.to_excel(output_path, index=False, engine='openpyxl')

        # Открываем файл для форматирования
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Определяем желтый цвет для подсветки
        yellow_fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type='solid')

        # Применяем подсветку
        for idx, row in df.iterrows():
            if row.get('_highlight', False):
                # Индекс строки в Excel (учитываем заголовок)
                excel_row_idx = idx + 2  # +1 для заголовка, +1 для индексации с 0

                # Подсвечиваем ячейку в колонке Departament (колонка A, индекс 1)
                cell = ws.cell(row=excel_row_idx, column=1)
                cell.fill = yellow_fill

        # Сохраняем
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Ошибка при экспорте: {e}")
        return False
